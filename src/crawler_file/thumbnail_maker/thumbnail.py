import fitz  # PyMuPDF
from PIL import Image, ImageDraw, ImageFont, ImageFile
from docx import Document
import aspose.slides as slides
import aspose.pydrawing as drawing
from moviepy.editor import VideoFileClip
import pandas as pd
import os  # Import os for file operations
import csv
from openpyxl import load_workbook
Image.MAX_IMAGE_PIXELS = None
ImageFile.LOAD_TRUNCATED_IMAGES = True 

class ThumbnailGenerator:
    def __init__(self, thumbnail_path, directory, http_url, thumbnail_type='PNG', thumbnail_size=(200, 200)):
        
        self.thumbnail_path = thumbnail_path
        self.thumbnail_size = thumbnail_size
        self.thumbnail_type = thumbnail_type
        self.thumbnail_directory = directory
        self.http_url = http_url

    def create_pdf_thumbnail(self, file, directory_path, page_num=0):
        pdf_document = fitz.open(file)
        
        # Select the specified page
        page = pdf_document.load_page(page_num)
        
        # Get a pixmap (image) of the page
        pix = page.get_pixmap()
        
        # Convert pixmap to a PIL image
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        
        # Resize image to the desired thumbnail size while maintaining aspect ratio
        img.thumbnail(self.thumbnail_size)  # Ensure the image fits within the thumbnail size
        
        # Save the thumbnail image
        img.save(self.make_thumbnail_directory(file, directory_path), self.thumbnail_type)

    def create_docx_thumbnail(self, file, directory_path, page_num=0):
        doc = Document(file)
        
        # Create a blank white image
        thumbnail = Image.new('RGB', (800, 600), 'white')
        draw = ImageDraw.Draw(thumbnail)
        
        # Extract and draw text on the thumbnail, limited to the specified number of paragraphs
        y_offset = 10
        paragraphs = doc.paragraphs[page_num:page_num + 5]  # Take 5 paragraphs starting from page_num
        for para in paragraphs:
            draw.text((10, y_offset), para.text, fill='black')
            y_offset += 20
        
        # Resize to thumbnail
        thumbnail.thumbnail(self.thumbnail_size)  # Ensure thumbnail size is applied
        
        # Save the thumbnail
        thumbnail.save(self.make_thumbnail_directory(file, directory_path), self.thumbnail_type)

    def create_pptx_thumbnail(self, file, directory_path, page_num=0):
        with slides.Presentation(file) as pres:
            # Ensure the page_num is within the valid range
            if page_num < 0 or page_num >= len(pres.slides):
                raise ValueError("page_num is out of range for the number of slides in the presentation.")

            # Select the specified slide
            slide = pres.slides[page_num]
            
            # Create a thumbnail image for the specified slide
            bmp = slide.get_thumbnail(1, 1)

            # Save the bitmap to a temporary file
            temp_file = f"{self.thumbnail_path}/temp_slide_thumbnail.png"
            bmp.save(temp_file, drawing.imaging.ImageFormat.png)

            # Open the temporary file as a PIL image
            img = Image.open(temp_file)

            # Resize the image to the desired thumbnail size while maintaining aspect ratio
            img.thumbnail(self.thumbnail_size)

            # Save the image to disk in the specified format
            img.save(self.make_thumbnail_directory(file, directory_path), self.thumbnail_type)

            # Optionally, remove the temporary file if you don't need it anymore
            os.remove(temp_file)

    def create_video_thumbnail(self, file, directory_path, t=1):
        video = VideoFileClip(file)
        
        # Get a frame at the specified time (in seconds)
        frame = video.get_frame(t)  # t = time in seconds where the frame is taken
        
        # Save the frame as an image
        frame_image = Image.fromarray(frame)
        
        # Resize the image to the desired thumbnail size while maintaining aspect ratio
        frame_image.thumbnail(self.thumbnail_size)

        # Save the thumbnail
        frame_image.save(self.make_thumbnail_directory(file, directory_path), self.thumbnail_type)

    def create_xlsx_thumbnail(self, file, directory_path):
        # Create a blank white image
        thumbnail = Image.new('RGB', (800, 600), 'white')
        draw = ImageDraw.Draw(thumbnail)
        
        # Load the workbook and the first sheet
        try:
            wb = load_workbook(file)
            sheet = wb.active
            
            # Read the first 5 rows
            rows = [sheet.iter_rows(min_row=i, max_row=i, values_only=True) for i in range(1, 6)]
        except Exception as e:
            print(f"Error reading {file}: {e}")
            return
        
        # Draw the text on the thumbnail
        y_offset = 10
        for row in rows:
            row_data = [str(cell) if cell is not None else '' for cell in next(row)]
            draw.text((10, y_offset), ', '.join(row_data), fill='black')
            y_offset += 20
        
        # Resize to thumbnail
        thumbnail.thumbnail(self.thumbnail_size)
        
        # Save the thumbnail
        thumbnail.save(self.make_thumbnail_directory(file, directory_path), self.thumbnail_type)

    def create_csv_thumbnail(self, file, directory_path):
        # Create a blank white image
        thumbnail = Image.new('RGB', (800, 600), 'white')
        draw = ImageDraw.Draw(thumbnail)
        
        # Read the CSV file
        try:
            with open(file, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                rows = [next(reader) for _ in range(5)]  # Get the first 5 rows
        except Exception as e:
            print(f"Error reading {file}: {e}")
            return
        
        # Draw the text on the thumbnail
        y_offset = 10
        for row in rows:
            draw.text((10, y_offset), ', '.join(row), fill='black')
            y_offset += 20
        
        # Resize to thumbnail
        thumbnail.thumbnail(self.thumbnail_size)
        
        # Save the thumbnail
        thumbnail.save(self.make_thumbnail_directory(file, directory_path), self.thumbnail_type)

    def create_txt_thumbnail(self, file, directory_path):
        txt_path = file

        # Create a blank white image
        thumbnail = Image.new('RGB', (800, 600), 'white')
        draw = ImageDraw.Draw(thumbnail)

        # Read the text file
        with open(txt_path, 'r') as file:
            lines = file.readlines()

        # Draw the text on the thumbnail (first 5 lines)
        y_offset = 10
        for line in lines[:5]:  # Limit to the first 5 lines
            draw.text((10, y_offset), line.strip(), fill='black')
            y_offset += 20

        # Resize to thumbnail
        thumbnail.thumbnail(self.thumbnail_size)

        # Save the thumbnail
        thumbnail.save(self.make_thumbnail_directory(file, directory_path), self.thumbnail_type)

    def create_md_thumbnail(self, file, directory_path):
        md_path = file

        # Create a blank white image
        thumbnail = Image.new('RGB', (800, 600), 'white')
        draw = ImageDraw.Draw(thumbnail)

        # Read the Markdown file with utf-8 encoding
        try:
            with open(md_path, 'r', encoding='utf-8') as file:
                lines = file.readlines()
        except UnicodeDecodeError as e:
            print(f"Error reading {file}: {e}")
            return

        # Draw the text on the thumbnail (first 5 lines)
        y_offset = 10
        for line in lines[:5]:  # Limit to the first 5 lines
            draw.text((10, y_offset), line.strip(), fill='black')
            y_offset += 20

        # Resize to thumbnail
        thumbnail.thumbnail(self.thumbnail_size)

        # Save the thumbnail
        thumbnail.save(self.make_thumbnail_directory(file, directory_path), self.thumbnail_type)

    def create_image_thumbnail(self, file, directory_path):
        
        try:
            # Disable the limit for maximum image size
            Image.MAX_IMAGE_PIXELS = None
            
            img = Image.open(file)
            
            # Resize the image to the desired thumbnail size while maintaining aspect ratio
            img.thumbnail(self.thumbnail_size)

            # Save the thumbnail
            img.save(self.make_thumbnail_directory(file, directory_path), self.thumbnail_type)
        
        except Exception as e:
            print(f"Error processing {file}: {e}")
    
    def ensure_directory_exists(self, path):
        if not os.path.exists(path):
            os.makedirs(path)
    
    def make_thumbnail_directory(self, file_path, directory_path):
        
        relative_path = os.path.relpath(str(file_path), directory_path)
        thumbnail_directory = os.path.dirname(relative_path)
        thumbnail_path = os.path.join(self.thumbnail_path, thumbnail_directory)
        self.ensure_directory_exists(thumbnail_path)
        return os.path.join(thumbnail_path, f"{os.path.basename(file_path)}.{self.thumbnail_type.lower()}")
    def make_thumbnail_relative_directory(self, file_path, directory_path):
        relative_path = os.path.relpath(str(file_path), directory_path)
        thumbnail_directory = os.path.dirname(relative_path)
        thumbnail_path = os.path.join(self.thumbnail_path, thumbnail_directory)
        self.ensure_directory_exists(thumbnail_path)
        path = f"{relative_path}.{self.thumbnail_type.lower()}"
        return path.replace('\\', '/')

    def create_thumbnail_from_path(self, file_path, directory_path):
        # Identify the file type based on the extension
        _, file_extension = os.path.splitext(file_path)
        file_extension = file_extension.lower()
        save_directory = self.make_thumbnail_directory(file_path, directory_path)

        # Call the appropriate thumbnail creation method based on the file type
        if file_extension == '.pdf':
            self.create_pdf_thumbnail(file_path, directory_path)
        elif file_extension == '.docx':
            self.create_docx_thumbnail(file_path, directory_path)
        elif file_extension in ['.pptx', '.ppt']:
            self.create_pptx_thumbnail(file_path, directory_path)
        elif file_extension in ['.mp4', '.avi', '.mov']:
            self.create_video_thumbnail(file_path, directory_path)
        elif file_extension in ['.xlsx', '.xls']:
            self.create_xlsx_thumbnail(file_path, directory_path)
        elif file_extension == '.csv':
            self.create_csv_thumbnail(file_path, directory_path)
        # elif file_extension == '.txt':
        #     self.create_txt_thumbnail(file_path, directory_path)
        elif file_extension == '.md':
            self.create_md_thumbnail(file_path, directory_path)
        elif file_extension in ['.jpg', '.jpeg', '.png', '.gif']:
            self.create_image_thumbnail(file_path, directory_path)
        else:
            print(f"Unsupported file type: {file_extension}")

    def delete_thumbnail_from_path(self, file_path, directory_path):
        save_directory = self.make_thumbnail_directory(file_path, directory_path)
        try:
            # Check if the file exists
            if os.path.exists(save_directory):
                os.remove(save_directory)
                print(f"Successfully deleted: {save_directory}")
            else:
                print(f"The file does not exist: {save_directory}")
        except Exception as e:
            print(f"Error deleting file: {e}")

    def crawl_directory_and_create_thumbnails(self, directory_path):
        if not os.path.isdir(directory_path):
            print(f"{directory_path} is not a valid directory.")
            return
        
        for root, dirs, files in os.walk(directory_path):
            for file in files:
                file_path = os.path.join(root, file)
                file_path.replace('\\', '/')
                print(f"Processing makeing thumbnails: {file_path}")
                self.create_thumbnail_from_path(file_path, directory_path)

# Example usage
# # Create thumbnails using the new method
# thumbnail_gen = ThumbnailGenerator(r"F:\______Restult\ElaticSearch\asdf\\")
# thumbnail_gen.crawl_directory_and_create_thumbnails(r"D:\PUG")
# print(thumbnail_gen.make_thumbnail_directory(r"F:\______Restult\ElaticSearch\thumbnail\sample.pdf", r"F:\______Restult\ElaticSearch"))
# thumbnail_gen.create_thumbnail_from_path(r"F:\______Restult\ElaticSearch\thumbnail\sample.pdf")
# thumbnail_gen.create_thumbnail_from_path(r"F:\______Restult\ElaticSearch\thumbnail\sample.docx")
# thumbnail_gen.create_thumbnail_from_path(r"F:\______Restult\ElaticSearch\thumbnail\sample.pptx")
# thumbnail_gen.create_thumbnail_from_path(r"F:\______Restult\ElaticSearch\thumbnail\sample.avi")
# thumbnail_gen.create_thumbnail_from_path(r"F:\______Restult\ElaticSearch\thumbnail\sample.xlsx")
# thumbnail_gen.create_thumbnail_from_path(r"F:\______Restult\ElaticSearch\thumbnail\sample.csv")
# thumbnail_gen.create_thumbnail_from_path(r"F:\______Restult\ElaticSearch\thumbnail\sample.txt")
# thumbnail_gen.create_thumbnail_from_path(r"F:\______Restult\ElaticSearch\thumbnail\sample.md")
# thumbnail_gen.create_thumbnail_from_path(r"F:\______Restult\ElaticSearch\thumbnail\sample.jpg")
# thumbnail_gen.crawl_directory_and_create_thumbnails(r"F:\______Restult\ElaticSearch\thumbnail")
